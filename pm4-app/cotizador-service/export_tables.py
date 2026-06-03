#!/usr/bin/env python3
"""
export_tables.py — se ejecuta UNA VEZ en build time.
Lee el Excel con openpyxl (data_only=True) y exporta todas las tablas
de lookup a tables.json para que app.py las use sin necesitar scipy.
"""
import json, os, sys
import openpyxl

HERE     = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.normpath(os.path.join(HERE, '..', 'frontend', 'src', 'resources', 'cotizador2.xlsx'))
OUTPUT   = os.path.join(HERE, 'tables.json')

print(f'Leyendo {TEMPLATE}…')
wb = openpyxl.load_workbook(TEMPLATE, data_only=True)

tables = {}

# ── D&O ───────────────────────────────────────────────────────────────────────
ws = wb["Configuraciond&o"]

# Límites disponibles (fila 34, cols D-P)
dyo_limites = []
for col in range(4, 17):
    v = ws.cell(34, col).value
    if v is not None:
        dyo_limites.append(v)

# Matriz de primas: fila 35-43, col C = facturación label, cols D-P = primas
dyo_matrix = []
for row in range(35, 44):
    fac = ws.cell(row, 3).value
    if fac is None:
        continue
    primas = []
    for col in range(4, 4 + len(dyo_limites)):
        v = ws.cell(row, col).value
        primas.append(None if (v is None or str(v) == 'N/A') else float(v))
    dyo_matrix.append({'fac': fac, 'primas': primas})

# NC/Anexo: fila 17-29, col B=límite, col C=límite_nc, col D=deducible
dyo_nc = {}
for row in range(17, 30):
    lim = ws.cell(row, 2).value
    lim_nc = ws.cell(row, 3).value
    ded    = ws.cell(row, 4).value
    if lim is not None:
        dyo_nc[str(int(lim))] = {
            'limite_nc': int(lim_nc) if lim_nc else None,
            'deducible': int(ded)    if ded    else None,
        }

tables['dyo'] = {
    'limites': [int(x) for x in dyo_limites],
    'matrix': dyo_matrix,
    'nc': dyo_nc,
}
print(f'  D&O: {len(dyo_matrix)} filas × {len(dyo_limites)} límites')

# ── CC / Infidelidad ──────────────────────────────────────────────────────────
ws_cc = wb['Nuevas Primas CYBER']

cc_primas = {}
for row in range(42, 600):
    k = ws_cc.cell(row, 5).value
    v = ws_cc.cell(row, 6).value
    if k is None:
        break
    if v is not None and str(v) not in ('N/A', '-', ''):
        cc_primas[str(k)] = float(v)

cc_ded = {}
for row in range(4, 12):
    k = ws_cc.cell(row, 22).value
    v = ws_cc.cell(row, 23).value
    if k is None:
        break
    cc_ded[str(k)] = str(v) if v is not None else None

tables['cc'] = {'primas': cc_primas, 'deducibles': cc_ded}
print(f'  CC: {len(cc_primas)} primas, {len(cc_ded)} deducibles')

# ── PDySI / Cyber (v2: ENTRADAS B27-B30, SALIDAS R23-R25) ────────────────────
ws_cy = wb['PRIMAS']

cy_primas = {}
for row in range(18, 60):
    k = ws_cy.cell(row, 1).value
    v = ws_cy.cell(row, 4).value
    if k is None:
        break
    if v is not None and str(v) not in ('N/A', '-', ''):
        try: cy_primas[str(k)] = float(v)
        except: pass

cy_ded = {}
for row in range(57, 100):
    k = ws_cy.cell(row, 1).value
    v = ws_cy.cell(row, 4).value
    if k is None:
        break
    if v is not None and str(v) not in ('N/A', '-', ''):
        try: cy_ded[str(k)] = float(v)
        except: pass

tables['pdysi'] = {'primas': cy_primas, 'deducibles': cy_ded}
print(f'  PDySI: {len(cy_primas)} primas, {len(cy_ded)} deducibles')

# ── PI (v2: 3 alternativas, key = {fac}{deducible}{limite}) ───────────────────
pi_tables = {}
for sector_name, sheet_name in [('ABOGADOS', 'ABOGADOS'), ('ADMINISTRADORES', 'ADMIN PH'), ('CONTADORES', 'CONTADORES')]:
    ws_pi = wb[sheet_name]
    lookup = {}
    for row in range(1, 800):
        key   = ws_pi.cell(row, 6).value   # col F = llave {fac}{ded}{lim}
        prima = ws_pi.cell(row, 5).value   # col E = prima
        if key and prima and str(prima) not in ('N/A', '-', 'PRIMA'):
            try: lookup[str(key)] = float(prima)
            except: pass
    pi_tables[sector_name] = lookup
    print(f'  PI/{sector_name}: {len(lookup)} entradas')

tables['pi'] = pi_tables

# Mapeo limit→deducible por sector (para auto-derivar deducible cuando no se pasa)
# Formato: {sector: {limite_str: deducible}}
pi_ded_map = {}
for sector_name, sheet_name in [('ABOGADOS', 'ABOGADOS'), ('ADMINISTRADORES', 'ADMIN PH'), ('CONTADORES', 'CONTADORES')]:
    ws_pi = wb[sheet_name]
    mapping = {}
    for row in range(1, 800):
        key   = ws_pi.cell(row, 6).value   # llave
        prima = ws_pi.cell(row, 5).value
        ded   = ws_pi.cell(row, 3).value   # col C = deducible
        lim   = ws_pi.cell(row, 4).value   # col D = limite
        if key and prima and ded and lim:
            try:
                lim_str = str(int(lim))
                if lim_str not in mapping:   # primer deducible encontrado para ese límite
                    mapping[lim_str] = int(ded)
            except: pass
    pi_ded_map[sector_name] = mapping
    print(f'  PI deducibles/{sector_name}: {len(mapping)} entradas')

tables['pi_ded_map'] = pi_ded_map

# ── Guardar ───────────────────────────────────────────────────────────────────
with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(tables, f, ensure_ascii=False)

print(f'\nOK Exportado a {OUTPUT} ({os.path.getsize(OUTPUT) // 1024} KB)')
