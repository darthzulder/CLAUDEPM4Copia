#!/usr/bin/env python3
import json
import os
import sys
import uuid
import argparse
import openpyxl
import requests
import subprocess
import re

# Reconfigure stdout to use UTF-8 to prevent console encoding crashes on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Global mappings for catalog names to mock records
MOCK_RECORDS = {
    "cat-canal": [
        {"codigo": "5", "descripcion": "Centro atención telefónica"},
        {"codigo": "3", "descripcion": "Web"},
        {"codigo": "1", "descripcion": "Presencial"},
        {"codigo": "2", "descripcion": "Virtual"},
        {"codigo": "4", "descripcion": "Escrito"}
    ],
    "cat-tipo-id": [
        {"codigo": "CC", "descripcion": "Cédula de Ciudadanía"},
        {"codigo": "CE", "descripcion": "Cédula de Extranjería"},
        {"codigo": "NIT", "descripcion": "NIT"},
        {"codigo": "Pasaporte", "descripcion": "Pasaporte"},
        {"codigo": "TI", "descripcion": "Tarjeta de Identidad"}
    ],
    "cat-tipo-persona": [
        {"codigo": "1", "descripcion": "Natural"},
        {"codigo": "2", "descripcion": "Jurídica"}
    ],
    "cat-pais": [
        {"codigo": "170", "descripcion": "Colombia"},
        {"codigo": "840", "descripcion": "Estados Unidos"},
        {"codigo": "724", "descripcion": "España"},
        {"codigo": "032", "descripcion": "Argentina"},
        {"codigo": "484", "descripcion": "México"}
    ],
    "cat-dpto": [
        {"codigo_departamento": "ANTIOQUIA", "nombre_departamento": "Antioquia"},
        {"codigo_departamento": "ATLANTICO", "nombre_departamento": "Atlántico"},
        {"codigo_departamento": "BOGOTA", "nombre_departamento": "Bogotá D.C."},
        {"codigo_departamento": "BOLIVAR", "nombre_departamento": "Bolívar"},
        {"codigo_departamento": "CUNDINAMARCA", "nombre_departamento": "Cundinamarca"},
        {"codigo_departamento": "MAGDALENA", "nombre_departamento": "Magdalena"},
        {"codigo_departamento": "RISARALDA", "nombre_departamento": "Risaralda"},
        {"codigo_departamento": "SANTANDER", "nombre_departamento": "Santander"},
        {"codigo_departamento": "VALLE", "nombre_departamento": "Valle del Cauca"}
    ],
    "cat-mpio": [
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "MEDELLIN", "nombre_municipio": "Medellín"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "BELLO", "nombre_municipio": "Bello"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "ENVIGADO", "nombre_municipio": "Envigado"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "ITAGUI", "nombre_municipio": "Itagüí"},
        {"codigo_departamento": "ANTIOQUIA", "codigo_municipio": "RIONEGRO", "nombre_municipio": "Rionegro"},
        {"codigo_departamento": "ATLANTICO", "codigo_municipio": "BARRANQUILLA", "nombre_municipio": "Barranquilla"},
        {"codigo_departamento": "ATLANTICO", "codigo_municipio": "SOLEDAD", "nombre_municipio": "Soledad"},
        {"codigo_departamento": "ATLANTICO", "codigo_municipio": "MALAMBO", "nombre_municipio": "Malambo"},
        {"codigo_departamento": "BOGOTA", "codigo_municipio": "BOGOTA", "nombre_municipio": "Bogotá D.C."},
        {"codigo_departamento": "BOLIVAR", "codigo_municipio": "CARTAGENA", "nombre_municipio": "Cartagena"},
        {"codigo_departamento": "BOLIVAR", "codigo_municipio": "TURBACO", "nombre_municipio": "Turbaco"},
        {"codigo_departamento": "BOLIVAR", "codigo_municipio": "MAGANGUE", "nombre_municipio": "Magangué"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "SOACHA", "nombre_municipio": "Soacha"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "CHIA", "nombre_municipio": "Chía"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "CAJICA", "nombre_municipio": "Cajicá"},
        {"codigo_departamento": "CUNDINAMARCA", "codigo_municipio": "FACATATIVA", "nombre_municipio": "Facatativá"},
        {"codigo_departamento": "MAGDALENA", "codigo_municipio": "SANTA_MARTA", "nombre_municipio": "Santa Marta"},
        {"codigo_departamento": "MAGDALENA", "codigo_municipio": "CIENAGA", "nombre_municipio": "Ciénaga"},
        {"codigo_departamento": "MAGDALENA", "codigo_municipio": "FUNDACION", "nombre_municipio": "Fundación"},
        {"codigo_departamento": "RISARALDA", "codigo_municipio": "PEREIRA", "nombre_municipio": "Pereira"},
        {"codigo_departamento": "RISARALDA", "codigo_municipio": "DOSQUEBRADAS", "nombre_municipio": "Dosquebradas"},
        {"codigo_departamento": "RISARALDA", "codigo_municipio": "SANTA_ROSA", "nombre_municipio": "Santa Rosa de Cabal"},
        {"codigo_departamento": "SANTANDER", "codigo_municipio": "BUCARAMANGA", "nombre_municipio": "Bucaramanga"},
        {"codigo_departamento": "SANTANDER", "codigo_municipio": "FLORIDABLANCA", "nombre_municipio": "Floridablanca"},
        {"codigo_departamento": "SANTANDER", "codigo_municipio": "GIRON", "nombre_municipio": "Girón"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "CALI", "nombre_municipio": "Cali"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "PALMIRA", "nombre_municipio": "Palmira"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "BUENAVENTURA", "nombre_municipio": "Buenaventura"},
        {"codigo_departamento": "VALLE", "codigo_municipio": "TULUA", "nombre_municipio": "Tuluá"}
    ],
    "cat-producto-sfc": [
        {"codigo_producto_sfc": "101", "nombre_producto_sfc": "Automóviles"},
        {"codigo_producto_sfc": "102", "nombre_producto_sfc": "Vida"},
        {"codigo_producto_sfc": "103", "nombre_producto_sfc": "Hogar"},
        {"codigo_producto_sfc": "104", "nombre_producto_sfc": "Salud"},
        {"codigo_producto_sfc": "105", "nombre_producto_sfc": "Cumplimiento"}
    ],
    "cat-motivo-sfc": [
        {"codigo": "301", "descripcion": "No pago siniestro"},
        {"codigo": "302", "descripcion": "Demora pago"},
        {"codigo": "303", "descripcion": "Mala atención"},
        {"codigo": "304", "descripcion": "Cobro indebido"},
        {"codigo": "305", "descripcion": "Incumplimiento de contrato"}
    ],
    "cat-tipo-sol": [
        {"codigo": "QUEJA_DIRECTA", "descripcion": "Queja Directa SmartSupervision"},
        {"codigo": "REQUERIMIENTO", "descripcion": "Requerimiento"},
        {"codigo": "SUGERENCIA", "descripcion": "Sugerencia"},
        {"codigo": "FELICITACION", "descripcion": "Felicitación"}
    ],
    "cat-instancia": [
        {"codigo": "2", "descripcion": "Entidad vigilada"},
        {"codigo": "1", "descripcion": "Defensor CF"},
        {"codigo": "3", "descripcion": "SFC"}
    ],
    "cat-punto": [
        {"codigo": "5", "descripcion": "Call Center"},
        {"codigo": "1", "descripcion": "Presencial"},
        {"codigo": "2", "descripcion": "Virtual"},
        {"codigo": "3", "descripcion": "Escrito"}
    ],
    "cat-admision": [
        {"codigo": "9", "descripcion": "No aplica"},
        {"codigo": "1", "descripcion": "Admitida"},
        {"codigo": "2", "descripcion": "No admitida"}
    ],
    "cat-ente": [
        {"codigo": "99", "descripcion": "Otros"},
        {"codigo": "1", "descripcion": "SFC"},
        {"codigo": "2", "descripcion": "Defensor CF"}
    ],
    "cat-sexo": [
        {"codigo": "M", "descripcion": "Masculino"},
        {"codigo": "F", "descripcion": "Femenino"},
        {"codigo": "I", "descripcion": "No informa"}
    ],
    "cat-lgbtiq": [
        {"codigo": "SI", "descripcion": "Pertenece a la comunidad LGBTIQ+"},
        {"codigo": "NO", "descripcion": "No pertenece"},
        {"codigo": "I", "descripcion": "No informa"}
    ],
    "cat-cond-esp": [
        {"codigo": "NINGUNA", "descripcion": "Ninguna"},
        {"codigo": "ADULTO_MAYOR", "descripcion": "Adulto mayor"},
        {"codigo": "DISCAPACIDAD_FISICA", "descripcion": "Discapacidad física"},
        {"codigo": "COGNITIVA", "descripcion": "Cognitiva"},
        {"codigo": "VULNERABLE", "descripcion": "Vulnerable"}
    ],
    "cat-prod-digital": [
        {"codigo": "1", "descripcion": "Sí"},
        {"codigo": "2", "descripcion": "No"}
    ],
    "cat-estado-queja": [
        {"codigo": "CERRADA_FAVOR_CF", "descripcion": "Cerrada a favor CF"},
        {"codigo": "CERRADA_FAVOR_ENTIDAD", "descripcion": "Cerrada a favor entidad"},
        {"codigo": "DESISTIDA", "descripcion": "Desistida"},
        {"codigo": "RECTIFICADA", "descripcion": "Rectificada"}
    ],
    "cat-favorab": [
        {"codigo": "1", "descripcion": "CF"},
        {"codigo": "2", "descripcion": "Entidad"},
        {"codigo": "3", "descripcion": "Parcial"}
    ],
    "cat-aceptacion": [
        {"codigo": "ACEPTADA_TOTAL", "descripcion": "Aceptada Total"},
        {"codigo": "ACEPTADA_PARCIAL", "descripcion": "Aceptada Parcial"},
        {"codigo": "RECHAZADA", "descripcion": "Rechazada"}
    ],
    "cat-rectif": [
        {"codigo": "RECTIFICADA", "descripcion": "Rectificada"},
        {"codigo": "NO_RECTIFICADA", "descripcion": "No Rectificada"}
    ],
    "cat-desist": [
        {"codigo": "DESISTIMIENTO_EXPRESO", "descripcion": "Desistimiento Expreso"},
        {"codigo": "DESISTIMIENTO_TACITO", "descripcion": "Desistimiento Tácito"}
    ],
    "cat-tutela": [
        {"codigo": "2", "descripcion": "No"},
        {"codigo": "1", "descripcion": "Sí"}
    ],
    "cat-marcacion": [
        {"codigo": "MARCADA", "descripcion": "Marcación especial"},
        {"codigo": "SIN_MARCACION", "descripcion": "Sin marcación"}
    ],
    "cat-expres": [
        {"codigo": "2", "descripcion": "No"},
        {"codigo": "1", "descripcion": "Sí"}
    ],
    "cat-tipo-fraude": [
        {"codigo": "FRAUDE_EXTERNO", "descripcion": "Fraude externo"},
        {"codigo": "FRAUDE_INTERNO", "descripcion": "Fraude interno"},
        {"codigo": "PHISHING", "descripcion": "Phishing"},
        {"codigo": "SUPLANTACION", "descripcion": "Suplantación"}
    ],
    "cat-mod-fraude": [
        {"codigo": "ROBO_INFO", "descripcion": "Robo de información"},
        {"codigo": "FALSIFICACION_DOCS", "descripcion": "Falsificación de documentos"},
        {"codigo": "SUPLANTACION_IDENTIDAD", "descripcion": "Suplantación de identidad"}
    ],
    "cat-area": [
        {"codigo_area": "SIN_AUTO", "nombre_area": "Siniestros Auto"},
        {"codigo_area": "SIN_VIDA", "nombre_area": "Siniestros Vida"},
        {"codigo_area": "PAGOS", "nombre_area": "Pagos"},
        {"codigo_area": "PRODUCTO", "nombre_area": "Producto"},
        {"codigo_area": "SAC", "nombre_area": "SAC"}
    ],
    "cat-usuarios-role": [
        {"codigo_area": "SIN_AUTO", "usuario": "juan.perez", "nombre_usuario": "Juan Pérez (Analista)", "rol": "Analista"},
        {"codigo_area": "SIN_AUTO", "usuario": "jorge.diaz", "nombre_usuario": "Jorge Díaz (Coordinador)", "rol": "Coordinador"},
        {"codigo_area": "SIN_VIDA", "usuario": "maria.gomez", "nombre_usuario": "María Gómez (Analista)", "rol": "Analista"},
        {"codigo_area": "PAGOS", "usuario": "carlos.rodriguez", "nombre_usuario": "Carlos Rodríguez (Coordinador)", "rol": "Coordinador"},
        {"codigo_area": "SAC", "usuario": "ana.martinez", "nombre_usuario": "Ana Martínez (Director)", "rol": "Director"},
        {"codigo_area": "PRODUCTO", "usuario": "luis.sanchez", "nombre_usuario": "Luis Sánchez (Coordinador)", "rol": "Coordinador"}
    ],
    "cat-motivo-reasig": [
        {"codigo": "ERROR_ASIGNACION", "descripcion": "Error asignación inicial"},
        {"codigo": "AREA_EQUIVOCADA", "descripcion": "Área equivocada"},
        {"codigo": "DERIVACION_PRODUCTO", "descripcion": "Derivación producto"}
    ],
    "cat-motivo-prorr": [
        {"codigo": "COMPLEJIDAD", "descripcion": "Complejidad del caso"},
        {"codigo": "FALTA_DOCUMENTACION", "descripcion": "Falta documentación del cliente"},
        {"codigo": "PROCESO_JUDICIAL", "descripcion": "Espera de proceso judicial"}
    ],
    "cat-rol-radicador": [
        {"codigo_instancia": "2", "codigo_rol_radicador": "CLIENTE", "nombre_rol_radicador": "Cliente"},
        {"codigo_instancia": "2", "codigo_rol_radicador": "INTERMEDIARIO", "nombre_rol_radicador": "Intermediario"},
        {"codigo_instancia": "2", "codigo_rol_radicador": "EMPLEADO", "nombre_rol_radicador": "Empleado Zurich"},
        {"codigo_instancia": "1", "codigo_rol_radicador": "DEFENSOR", "nombre_rol_radicador": "Defensor del Consumidor"}
    ],
    "cat-tipo-solic-pqrs": [
        {"codigo": "SOLICITUD", "descripcion": "Solicitud"},
        {"codigo": "FELICITACION", "descripcion": "Felicitación"},
        {"codigo": "QUEJA", "descripcion": "Queja"},
        {"codigo": "SUGERENCIA", "descripcion": "Sugerencia"},
        {"codigo": "PETICION", "descripcion": "Derecho de petición"}
    ],
    "cat-detalle-producto": [
        {"codigo_producto_sfc": "101", "codigo_detalle_producto": "10101", "nombre_detalle_producto": "Autos Livianos"},
        {"codigo_producto_sfc": "101", "codigo_detalle_producto": "10102", "nombre_detalle_producto": "Autos Pesados"},
        {"codigo_producto_sfc": "102", "codigo_detalle_producto": "10201", "nombre_detalle_producto": "Vida Individual"},
        {"codigo_producto_sfc": "102", "codigo_detalle_producto": "10202", "nombre_detalle_producto": "Vida Colectivo"},
        {"codigo_producto_sfc": "103", "codigo_detalle_producto": "10301", "nombre_detalle_producto": "Incendio Hogar"},
        {"codigo_producto_sfc": "103", "codigo_detalle_producto": "10302", "nombre_detalle_producto": "Terremoto Hogar"}
    ]
}


def decrypt_token(blob, key_raw):
    if not blob or blob.startswith("eyJ"):
        return blob
    
    # Node code for decryption
    node_code = f"""
const crypto = require('crypto');
try {{
    const key = crypto.createHash('sha256').update({json.dumps(key_raw)}).digest();
    const buf = Buffer.from({json.dumps(blob)}.replace(/-/g, '+').replace(/_/g, '/'), 'base64');
    const iv = buf.subarray(0, 16);
    const ciphertext = buf.subarray(16);
    const decipher = crypto.createDecipheriv('aes-256-cbc', key, iv);
    const decrypted = Buffer.concat([decipher.update(ciphertext), decipher.final()]);
    const payload = JSON.parse(decrypted.toString('utf8'));
    console.log(payload.token);
}} catch (e) {{
    console.error(e.message);
    process.exit(1);
}}
"""
    # Try running local node first
    try:
        res = subprocess.run(["node", "-e", node_code], capture_output=True, text=True, check=True)
        return res.stdout.strip()
    except (subprocess.SubprocessError, FileNotFoundError):
        # If local node fails or is not found, try via docker
        try:
            res = subprocess.run(["docker", "exec", "pm4-app-container", "node", "-e", node_code], capture_output=True, text=True, check=True)
            return res.stdout.strip()
        except Exception as e:
            print(f"Error al desencriptar el token de PM4: {e}")
            print("Asegúrate de que 'node' local o el contenedor docker 'pm4-app-container' estén disponibles.")
            sys.exit(1)


def load_env(env_path):
    env_vars = {}
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    env_vars[k.strip()] = v.strip()
    return env_vars


def get_catalog_fields(slug):
    """Retorna los campos de base de datos / inputs de formulario requeridos según el slug."""
    if slug == "cat-dpto":
        return [
            {"name": "codigo_departamento", "label": "Código Departamento"},
            {"name": "nombre_departamento", "label": "Nombre Departamento"}
        ]
    elif slug == "cat-mpio":
        return [
            {"name": "codigo_departamento", "label": "Código Departamento (Padre)"},
            {"name": "codigo_municipio", "label": "Código Municipio"},
            {"name": "nombre_municipio", "label": "Nombre Municipio"}
        ]
    elif slug == "cat-area":
        return [
            {"name": "codigo_area", "label": "Código Área"},
            {"name": "nombre_area", "label": "Nombre Área"}
        ]
    elif slug == "cat-usuarios-role":
        return [
            {"name": "codigo_area", "label": "Código Área (Padre)"},
            {"name": "usuario", "label": "Usuario"},
            {"name": "nombre_usuario", "label": "Nombre de Usuario"},
            {"name": "rol", "label": "Rol"}
        ]
    elif slug == "cat-producto-sfc":
        return [
            {"name": "codigo_producto_sfc", "label": "Código Producto SFC"},
            {"name": "nombre_producto_sfc", "label": "Nombre Producto SFC"}
        ]
    elif slug == "cat-detalle-producto":
        return [
            {"name": "codigo_producto_sfc", "label": "Código Producto SFC (Padre)"},
            {"name": "codigo_detalle_producto", "label": "Código Detalle Producto"},
            {"name": "nombre_detalle_producto", "label": "Nombre Detalle Producto"}
        ]
    elif slug == "cat-rol-radicador":
        return [
            {"name": "codigo_instancia", "label": "Código Instancia (Padre)"},
            {"name": "codigo_rol_radicador", "label": "Código Rol Radicador"},
            {"name": "nombre_rol_radicador", "label": "Nombre Rol Radicador"}
        ]
    else:
        # Catálogo estándar (con código y descripción)
        return [
            {"name": "codigo", "label": "Código"},
            {"name": "descripcion", "label": "Descripción"}
        ]


def clean_screen_title(title):
    # Remueve cualquier caracter que no sea alfanumérico, espacio o guion
    return re.sub(r"[^a-zA-Z0-9\s\-]", "", title).strip()


def generate_screen_json(title, fields, is_view=False, screen_category_id="1"):
    """Genera la estructura JSON de una pantalla de ProcessMaker 4."""
    cleaned_title = clean_screen_title(title)
    items = []
    
    # 1. Inputs del Formulario
    for field in fields:
        items.append({
            "uuid": str(uuid.uuid4()),
            "label": "Line Input",
            "component": "FormInput",
            "config": {
                "icon": "far fa-square",
                "name": field["name"],
                "type": "text",
                "label": field["label"],
                "helper": None,
                "readonly": is_view,
                "dataFormat": "string",
                "validation": [] if is_view else ["required"],
                "placeholder": f"Ingrese {field['label'].lower()}",
                "defaultValue": { "mode": "js", "value": None },
                "conditionalHide": None,
                "customCssSelector": None
            },
            "editor-control": "FormInput",
            "editor-component": "FormInput"
        })
    
    # 2. Botón de guardar si no es pantalla de lectura
    if not is_view:
        items.append({
            "uuid": str(uuid.uuid4()),
            "label": "Submit Button",
            "component": "FormButton",
            "config": {
                "icon": "fas fa-share-square",
                "name": None,
                "event": "submit",
                "label": "GRABAR",
                "loading": False,
                "tooltip": [],
                "variant": "primary",
                "fieldValue": None,
                "loadingLabel": "Loading...",
                "defaultSubmit": True
            },
            "editor-control": "FormSubmit",
            "editor-component": "FormButton"
        })
        
    return {
        "title": f"Ver - {cleaned_title}" if is_view else f"Crear - {cleaned_title}",
        "description": f"Pantalla de visualización para {title}" if is_view else f"Pantalla de creación/edición para {title}",
        "type": "FORM",
        "screen_category_id": screen_category_id,
        "config": [
            {
                "name": title,
                "order": 1,
                "items": items
            }
        ],
        "computed": [],
        "watchers": [],
        "custom_css": None,
        "status": "ACTIVE",
        "key": None,
        "is_template": 0,
        "is_default": 0
    }


def create_screen(base_url, token, screen_data):
    """POST /screens en PM4 API"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/1.0/screens"
    payload = {
        "title": screen_data["title"],
        "type": screen_data["type"],
        "description": screen_data["description"],
        "config": screen_data["config"],
        "computed": screen_data.get("computed", []),
        "watchers": screen_data.get("watchers", []),
        "custom_css": screen_data.get("custom_css"),
        "screen_category_id": screen_data.get("screen_category_id")
    }
    r = requests.post(url, json=payload, headers=headers)
    r.raise_for_status()
    return r.json()


def get_existing_collections(base_url, token):
    """GET /collections para mapear slugs existentes a IDs"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/1.0/collections"
    r = requests.get(url, params={"per_page": 500}, headers=headers)
    r.raise_for_status()
    data = r.json()
    return {col["name"]: col["id"] for col in data.get("data", [])}


def create_collection(base_url, token, col_data):
    """POST /collections en PM4 API"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/1.0/collections"
    r = requests.post(url, json=col_data, headers=headers)
    r.raise_for_status()
    return r.json()


def truncate_collection(base_url, token, collection_id):
    """DELETE /collections/{id}/truncate para borrar registros viejos"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/1.0/collections/{collection_id}/truncate"
    r = requests.delete(url, headers=headers)
    r.raise_for_status()
    return r.status_code


def create_record(base_url, token, collection_id, record_data):
    """POST /collections/{id}/records en PM4 API"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/1.0/collections/{collection_id}/records"
    payload = {
        "data": record_data
    }
    r = requests.post(url, json=payload, headers=headers)
    r.raise_for_status()
    return r.json()


def get_screen_category_id(base_url, token):
    """Obtiene el ID de categoría de pantalla buscando 'Fast Flow - Zurich'"""
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    url = f"{base_url}/api/1.0/screen_categories"
    try:
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        cats = r.json().get("data", [])
        for c in cats:
            if c.get("name") == "Fast Flow - Zurich":
                return str(c.get("id"))
        if cats:
            return str(cats[0].get("id"))
    except Exception as e:
        print(f"Advertencia: No se pudo obtener las categorías de screens ({e}). Usando fallback ID '1'.")
    return "1"


def main():
    parser = argparse.ArgumentParser(description="Script para importar catálogos de excel a PM4.")
    parser.add_argument("--dry-run", action="store_true", help="Valida localmente la estructuración de catálogos y pantallas sin llamar a la API.")
    args = parser.parse_args()

    # Rutas relativas
    here = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(here, "insumos", "Anexo02_Mockups_TOBE_QuejaDirectas_v3_0.xlsx")
    env_path = os.path.join(here, ".env")

    print("=== PROCESO DE MIGRACIÓN DE CATÁLOGOS PM4 ===")
    print(f"Cargando Excel desde: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"Error: No se encontró el archivo Excel en {excel_path}")
        sys.exit(1)

    # 1. Cargar variables de entorno si no es Dry Run
    pm4_url = None
    pm4_token = None
    screen_category_id = "7" # Default para dry-run
    if not args.dry_run:
        print(f"Cargando variables desde: {env_path}")
        env = load_env(env_path)
        pm4_url = env.get("PM4_BASE_URL")
        pm4_token = env.get("PM4_TOKEN")
        
        # Decrypt token if it is encrypted
        pm4_token = decrypt_token(pm4_token, env.get("IFRAME_ENCRYPTION_KEY"))
        
        # Obtener el ID de categoría de pantalla dinámicamente
        screen_category_id = get_screen_category_id(pm4_url, pm4_token)
        print(f"Usando ID de categoría de Screen: {screen_category_id}")

        if not pm4_url or not pm4_token:
            print("Error: PM4_BASE_URL y PM4_TOKEN deben estar definidos en el archivo .env")
            sys.exit(1)
        print(f"PM4 base: {pm4_url}")

    # 2. Leer hoja de Catálogos
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "07_Catalogs" not in wb.sheetnames:
        print("Error: La hoja '07_Catalogs' no existe en el libro de Excel.")
        sys.exit(1)

    ws = wb["07_Catalogs"]
    catalogs = []
    
    # Empezamos en la fila 6
    row = 6
    while True:
        cat_id = ws.cell(row, 2).value
        nombre = ws.cell(row, 3).value
        if not cat_id and not nombre:
            # Fin del listado
            break
        
        if cat_id:
            slug = str(cat_id).strip().lower()
            catalogs.append({
                "id": str(cat_id).strip(),
                "slug": slug,
                "nombre": str(nombre).strip() if nombre else str(cat_id).strip(),
                "description": ws.cell(row, 5).value or ""
            })
        row += 1

    print(f"Se identificaron {len(catalogs)} catálogos en el Excel.")

    # 3. Consultar colecciones existentes si no es dry run
    existing_cols = {}
    if not args.dry_run:
        try:
            print("Obteniendo colecciones existentes de PM4…")
            existing_cols = get_existing_collections(pm4_url, pm4_token)
            print(f"Encontradas {len(existing_cols)} colecciones en la instancia.")
        except Exception as e:
            print(f"Error al conectar con la API de PM4: {e}")
            sys.exit(1)

    # 4. Iterar y procesar cada catálogo
    success_count = 0
    for idx, cat in enumerate(catalogs):
        slug = cat["slug"]
        nombre = cat["nombre"]
        desc = cat["description"]

        print(f"\n[{idx+1}/{len(catalogs)}] Procesando {cat['id']} ('{nombre}')…")
        
        # Obtener los campos
        fields = get_catalog_fields(slug)
        records = MOCK_RECORDS.get(slug, [])

        if not records:
            # Fallback generico por si falta mapeo
            print(f"  [!] Advertencia: Sin datos simulados definidos para {slug}. Se creará vacío.")
        
        print(f"  Campos definidos: {[f['name'] for f in fields]}")
        print(f"  Cantidad de registros sugeridos: {len(records)}")

        if args.dry_run:
            # Generar JSONs para validar sintaxis de Dry Run
            create_screen_json = generate_screen_json(nombre, fields, is_view=False, screen_category_id=screen_category_id)
            view_screen_json = generate_screen_json(nombre, fields, is_view=True, screen_category_id=screen_category_id)
            print(f"  [Dry Run] Generada Create Screen '{create_screen_json['title']}' con {len(create_screen_json['config'][0]['items'])} items.")
            print(f"  [Dry Run] Generada View Screen '{view_screen_json['title']}' con {len(view_screen_json['config'][0]['items'])} items.")
            success_count += 1
            continue

        # Proceso real llamando a API
        try:
            col_id = existing_cols.get(slug)
            
            if col_id:
                print(f"  La colección '{slug}' ya existe (ID: {col_id}). Limpiando registros antiguos…")
                truncate_collection(pm4_url, pm4_token, col_id)
                print("  Registros eliminados exitosamente.")
            else:
                # 1. Crear Screen de Creación/Edición
                print("  Creando pantalla de Creación/Edición…")
                create_screen_data = generate_screen_json(nombre, fields, is_view=False, screen_category_id=screen_category_id)
                c_screen = create_screen(pm4_url, pm4_token, create_screen_data)
                create_screen_id = c_screen["id"]
                print(f"    Creada con ID: {create_screen_id}")

                # 2. Crear Screen de Visualización (Readonly)
                print("  Creando pantalla de Visualización (Readonly)…")
                view_screen_data = generate_screen_json(nombre, fields, is_view=True, screen_category_id=screen_category_id)
                v_screen = create_screen(pm4_url, pm4_token, view_screen_data)
                read_screen_id = v_screen["id"]
                print(f"    Creada con ID: {read_screen_id}")

                # 3. Crear Colección
                print("  Creando la colección…")
                col_payload = {
                    "name": slug,
                    "custom_title": nombre,
                    "description": desc or f"Colección para {nombre}",
                    "create_screen_id": str(create_screen_id),
                    "read_screen_id": str(read_screen_id),
                    "update_screen_id": str(create_screen_id),
                    "signal_create": False,
                    "signal_update": False,
                    "signal_delete": False
                }
                new_col = create_collection(pm4_url, pm4_token, col_payload)
                col_id = new_col["id"]
                print(f"    Colección creada con ID: {col_id}")

            # 4. Insertar los Registros
            if records:
                print(f"  Insertando {len(records)} registros…")
                inserted = 0
                for rec in records:
                    create_record(pm4_url, pm4_token, col_id, rec)
                    inserted += 1
                print(f"    Importados {inserted}/{len(records)} registros con éxito.")

            success_count += 1
        except Exception as e:
            print(f"  [X] ERROR procesando catálogo {cat['id']}: {e}")
            if isinstance(e, requests.exceptions.HTTPError):
                print(f"    Detalle HTTP: {e.response.text}")

    print("\n==============================================")
    if args.dry_run:
        print(f"DRY RUN COMPLETADO. Se validaron {success_count}/{len(catalogs)} catálogos con éxito.")
    else:
        print(f"MIGRACIÓN COMPLETADA. Se procesaron {success_count}/{len(catalogs)} catálogos con éxito.")
    print("==============================================")


if __name__ == "__main__":
    main()
